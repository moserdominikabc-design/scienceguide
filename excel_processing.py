from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from factory_columns import ALIASES, FACTORY_COLUMNS, PREFER_LAST_DUPLICATE


@dataclass(frozen=True)
class SourceColumn:
    index: int  # 1-based Excel column index
    header: str
    display: str


def _text(value) -> str:
    return "" if value is None else str(value)


def extract_date_from_filename(path: str | Path) -> str | None:
    """Return YYYY_MM_DD if present in the filename."""
    match = re.search(r"(?<!\d)(20\d{2})[_-](\d{2})[_-](\d{2})(?!\d)", Path(path).name)
    if not match:
        return None
    return "_".join(match.groups())


def read_source_columns(path: str | Path) -> list[SourceColumn]:
    path = Path(path)
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Input exports must be .xlsx or .xlsm files.")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    finally:
        wb.close()

    counts: dict[str, int] = {}
    for header in headers:
        if header is not None:
            counts[_text(header)] = counts.get(_text(header), 0) + 1

    columns: list[SourceColumn] = []
    for index, raw_header in enumerate(headers, start=1):
        if raw_header is None or _text(raw_header).strip() == "":
            continue
        header = _text(raw_header)
        display = header
        if counts.get(header, 0) > 1:
            display = f"{header}  [source column {index}]"
        columns.append(SourceColumn(index=index, header=header, display=display))
    return columns


def _candidate_indices(factory_name: str, columns: Sequence[SourceColumn]) -> list[int]:
    exact = [c.index for c in columns if c.header == factory_name]
    if exact:
        return exact

    # Case-insensitive fallback handles the legacy ANG 'barrierefreiheit' spelling.
    folded = factory_name.casefold()
    casefold_matches = [c.index for c in columns if c.header.casefold() == folded]
    if casefold_matches:
        return casefold_matches

    for alias in ALIASES.get(factory_name, []):
        alias_matches = [c.index for c in columns if c.header == alias]
        if alias_matches:
            return alias_matches
        alias_folded = alias.casefold()
        alias_casefold = [c.index for c in columns if c.header.casefold() == alias_folded]
        if alias_casefold:
            return alias_casefold
    return []


def factory_selection(kind: str, columns: Sequence[SourceColumn]) -> tuple[list[int], list[str], dict[str, str]]:
    """Return selected source indices, missing factory fields, and alias matches."""
    kind = kind.upper()
    if kind not in FACTORY_COLUMNS:
        raise ValueError(f"Unknown export kind: {kind}")

    by_index = {c.index: c for c in columns}
    selected: list[int] = []
    missing: list[str] = []
    alias_matches: dict[str, str] = {}
    used: set[int] = set()

    for factory_name in FACTORY_COLUMNS[kind]:
        candidates = [i for i in _candidate_indices(factory_name, columns) if i not in used]
        if not candidates:
            missing.append(factory_name)
            continue

        if factory_name in PREFER_LAST_DUPLICATE.get(kind, set()):
            chosen = candidates[-1]
        else:
            chosen = candidates[0]
        selected.append(chosen)
        used.add(chosen)

        actual = by_index[chosen].header
        if actual != factory_name:
            alias_matches[factory_name] = actual

    return selected, missing, alias_matches


def _parse_changed_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None

    text = str(value).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%Y %H:%M", "%d.%m.%Y %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _find_header_index(headers: Sequence[object], wanted: str) -> int | None:
    wanted_folded = wanted.casefold()
    for idx, header in enumerate(headers):
        if header is not None and str(header).casefold() == wanted_folded:
            return idx
    return None


def _is_technical_mapping_row(row: Sequence[object], id_idx: int | None, changed_idx: int | None) -> bool:
    if id_idx is None or changed_idx is None:
        return False
    id_value = _text(row[id_idx]).strip().upper()
    changed_value = _text(row[changed_idx]).strip().upper()
    return id_value == "ID" and changed_value in {"GEAENDERT", "GEÄNDERT"}


def process_workbook(
    input_path: str | Path,
    selected_indices: Sequence[int],
    output_path: str | Path,
) -> dict[str, object]:
    """Process one export and write a clean workbook using selected source columns."""
    input_path = Path(input_path)
    output_path = Path(output_path)

    if not selected_indices:
        raise ValueError("At least one output column must be selected.")

    wb_in = load_workbook(input_path, read_only=True, data_only=True)
    try:
        ws_in = wb_in[wb_in.sheetnames[0]]
        header_row = [cell.value for cell in next(ws_in.iter_rows(min_row=1, max_row=1))]
        max_col = len(header_row)
        bad = [i for i in selected_indices if i < 1 or i > max_col]
        if bad:
            raise ValueError(f"Selected source column(s) outside workbook range: {bad}")

        changed_idx = _find_header_index(header_row, "Geändert am")
        id_idx = _find_header_index(header_row, "ID")

        rows = [tuple(row) for row in ws_in.iter_rows(min_row=2, values_only=True)]
    finally:
        wb_in.close()

    before = len(rows)
    rows = [row for row in rows if not _is_technical_mapping_row(row, id_idx, changed_idx)]
    mapping_rows_removed = before - len(rows)

    if changed_idx is not None:
        def sort_key(row: Sequence[object]):
            parsed = _parse_changed_date(row[changed_idx])
            return (parsed is not None, parsed or date.min)

        rows.sort(key=sort_key, reverse=True)

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Angebote"

    output_headers = [header_row[i - 1] for i in selected_indices]
    ws_out.append(output_headers)

    header_fill = PatternFill(fill_type="solid", fgColor="FFE7EEF8")
    for cell in ws_out[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    selected_zero = [i - 1 for i in selected_indices]
    changed_output_pos = None
    if changed_idx is not None and changed_idx in selected_zero:
        changed_output_pos = selected_zero.index(changed_idx) + 1

    for row in rows:
        values = [row[i] if i < len(row) else None for i in selected_zero]
        if changed_output_pos is not None:
            parsed = _parse_changed_date(values[changed_output_pos - 1])
            if parsed is not None:
                values[changed_output_pos - 1] = parsed
        ws_out.append(values)

    if changed_output_pos is not None:
        for cells in ws_out.iter_rows(min_row=2, min_col=changed_output_pos, max_col=changed_output_pos):
            cells[0].number_format = "DD.MM.YYYY"

    ws_out.freeze_panes = "A2"
    if ws_out.max_column and ws_out.max_row:
        ws_out.auto_filter.ref = ws_out.dimensions

    # Conservative widths: readable without creating extremely wide worksheets.
    for col_idx in range(1, ws_out.max_column + 1):
        header = _text(ws_out.cell(row=1, column=col_idx).value)
        ws_out.column_dimensions[ws_out.cell(row=1, column=col_idx).column_letter].width = min(max(len(header) + 2, 12), 38)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(output_path)
    wb_out.close()

    return {
        "input_rows": before,
        "output_rows": len(rows),
        "mapping_rows_removed": mapping_rows_removed,
        "output_columns": len(selected_indices),
        "output_path": str(output_path),
    }
