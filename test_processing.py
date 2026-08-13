from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from excel_processing import factory_selection, process_workbook, read_source_columns
from user_defaults import save_default, user_default_selection


class SourceDrivenColumnTests(unittest.TestCase):
    def make_workbook(self, path: Path, headers: list[str], rows: list[list[object]]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        wb.save(path)
        wb.close()

    def test_all_possible_columns_come_from_selected_file(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "2026_08_13_ver.xlsx"
            headers = ["ID", "Titel_DE", "Geändert am", "Brand New Future Field"]
            self.make_workbook(path, headers, [[1, "Example", "13.08.2026", "New value"]])

            columns = read_source_columns(path)
            self.assertEqual([c.header for c in columns], headers)

            selected, missing, aliases = factory_selection("VER", columns)
            self.assertIn(2, selected)  # Titel_DE factory default
            self.assertIn(3, selected)  # Geändert am factory default
            self.assertNotIn(4, selected)  # new source field is available, but not factory-selected
            self.assertIn("Brand New Future Field", [c.header for c in columns])
            self.assertTrue(missing)
            self.assertEqual(aliases, {})

    def test_unknown_source_column_can_be_exported(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            td = Path(td)
            source = td / "2026_08_13_ver.xlsx"
            output = td / "out.xlsx"
            self.make_workbook(
                source,
                ["ID", "Geändert am", "New Column"],
                [
                    ["ID", "GEAENDERT", "NEW_FIELD"],
                    [1, "12.08.2026", "older"],
                    [2, "13.08.2026", "newer"],
                ],
            )

            result = process_workbook(source, [2, 3], output)
            self.assertEqual(result["mapping_rows_removed"], 1)
            self.assertEqual(result["output_columns"], 2)

            wb = load_workbook(output, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            self.assertEqual([c.value for c in ws[1]], ["Geändert am", "New Column"])
            self.assertEqual(ws.cell(2, 2).value, "newer")
            self.assertEqual(ws.cell(3, 2).value, "older")
            wb.close()

    def test_duplicate_source_headers_remain_individually_selectable(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "ang.xlsx"
            self.make_workbook(
                path,
                ["Geändert am", "ÖV-Haltestelle", "ÖV-Haltestelle", "Extra"],
                [["13.08.2026", "general", "route", "x"]],
            )
            columns = read_source_columns(path)
            ov = [c for c in columns if c.header == "ÖV-Haltestelle"]
            self.assertEqual([c.index for c in ov], [2, 3])
            self.assertNotEqual(ov[0].display, ov[1].display)

    def test_single_kind_processing_does_not_depend_on_other_kind(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            td = Path(td)
            source = td / "2026_08_13_ang.xlsx"
            output = td / "2026_08_13_ang_mod.xlsx"
            self.make_workbook(
                source,
                ["ID", "Geändert am", "Titel_DE"],
                [
                    ["ID", "GEAENDERT", "TITEL_D"],
                    [1, "12.08.2026", "older"],
                    [2, "13.08.2026", "newer"],
                ],
            )

            result = process_workbook(source, [2, 3], output)
            self.assertTrue(output.exists())
            self.assertEqual(result["mapping_rows_removed"], 1)
            self.assertEqual(result["output_rows"], 2)

            wb = load_workbook(output, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            self.assertEqual(ws.cell(2, 2).value, "newer")
            self.assertEqual(ws.cell(3, 2).value, "older")
            wb.close()

    def test_personal_default_persists_by_header_and_duplicate_occurrence(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            td = Path(td)
            source_a = td / "a.xlsx"
            source_b = td / "b.xlsx"
            defaults_file = td / "column_defaults.json"

            self.make_workbook(
                source_a,
                ["ID", "ÖV-Haltestelle", "Titel_DE", "ÖV-Haltestelle", "Extra"],
                [[1, "general", "Title", "route", "x"]],
            )
            columns_a = read_source_columns(source_a)
            save_default("ANG", columns_a, [4, 3], path=defaults_file)

            # Insert a new source column before the saved fields. The stored default
            # must follow semantic column identity, not the old Excel positions.
            self.make_workbook(
                source_b,
                ["New", "ID", "ÖV-Haltestelle", "Titel_DE", "ÖV-Haltestelle", "Extra"],
                [["n", 1, "general", "Title", "route", "x"]],
            )
            columns_b = read_source_columns(source_b)
            selected, missing = user_default_selection("ANG", columns_b, path=defaults_file)
            self.assertEqual(selected, [5, 4])
            self.assertEqual(missing, [])

    def test_personal_default_reports_fields_missing_from_new_export(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            td = Path(td)
            source = td / "source.xlsx"
            changed = td / "changed.xlsx"
            defaults_file = td / "column_defaults.json"
            self.make_workbook(source, ["Titel_DE", "Extra"], [["A", "x"]])
            save_default("VER", read_source_columns(source), [1, 2], path=defaults_file)
            self.make_workbook(changed, ["Titel_DE"], [["A"]])

            selected, missing = user_default_selection("VER", read_source_columns(changed), path=defaults_file)
            self.assertEqual(selected, [1])
            self.assertEqual(missing, ["Extra"])


if __name__ == "__main__":
    unittest.main()
