from __future__ import annotations

import argparse
from pathlib import Path
import tempfile

from openpyxl import load_workbook

from excel_processing import factory_selection, process_workbook, read_source_columns
from factory_columns import FACTORY_COLUMNS


def main() -> None:
    parser = argparse.ArgumentParser(description="Verify ScienceGuide Cleaner against representative VER/ANG exports.")
    parser.add_argument("ver", type=Path, help="Path to a VER / Veranstaltungen .xlsx export")
    parser.add_argument("ang", type=Path, help="Path to an ANG / Angebote .xlsx export")
    args = parser.parse_args()

    ver_cols = read_source_columns(args.ver)
    ang_cols = read_source_columns(args.ang)
    ver_sel, ver_missing, ver_alias = factory_selection("VER", ver_cols)
    ang_sel, ang_missing, ang_alias = factory_selection("ANG", ang_cols)

    assert len(FACTORY_COLUMNS["VER"]) == 51
    assert len(FACTORY_COLUMNS["ANG"]) == 59

    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        ver_out = td / "ver.xlsx"
        ang_out = td / "ang.xlsx"
        vr = process_workbook(args.ver, ver_sel, ver_out)
        ar = process_workbook(args.ang, ang_sel, ang_out)

        for path in (ver_out, ang_out):
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            assert ws.max_row >= 1
            assert ws.max_column >= 1
            wb.close()

    print("Verification passed.")
    print("VER:", vr)
    print("  missing factory fields:", ver_missing)
    print("  renamed matches:", ver_alias)
    print("ANG:", ar)
    print("  missing factory fields:", ang_missing)
    print("  renamed matches:", ang_alias)


if __name__ == "__main__":
    main()
